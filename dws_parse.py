import os, json, pickle
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


def parse_dws_data():
    awqi_path = "scripts/dws_spreadsheet/data/AWQI - Raw Data 2020_21_EN.xlsx"
    ws = load_workbook(awqi_path)["AWQI-RAW"]

    name2_param_map = {}  # 4, 11, 12
    for idx, row in enumerate(ws.rows):
        if idx == 0:  # skip the header
            continue
        dws_name, param_name, param_group = row[3].value, row[10].value, row[11].value
        dws_name = dws_name.strip().lower()
        if dws_name not in name2_param_map:
            name2_param_map[dws_name] = {"param_name": set(), "param_group": set()}
        name2_param_map[dws_name]["param_name"].add(param_name.lower().capitalize())
        if param_group is None:  # sometimes it is empty
            continue
        name2_param_map[dws_name]["param_group"].add(param_group.lower().capitalize())

    # update the columns in the existing workbook
    compiled_wb_path = "scripts/dws_spreadsheet/data/compiled_dws_data.xlsx"
    wb = load_workbook(compiled_wb_path)
    ws = wb["Sheet1"]
    ws["E1"], ws["F1"] = "Paramater Name", "Parameter Group"
    n_rows = len(list(ws.rows))
    for idx, row in enumerate(ws.rows):
        if idx == 0:
            continue
        if idx == 1517:  # some empty last rows
            break
        dws_name = ws[f"A{idx+1}"].value.strip().lower().strip()
        if dws_name not in name2_param_map:
            continue

        ws[f"E{idx+1}"] = ", ".join(name2_param_map[dws_name]["param_name"])
        ws[f"F{idx+1}"] = ", ".join(name2_param_map[dws_name]["param_group"])
    wb.save("scripts/dws_spreadsheet/data/compiled_dws_datadone.xlsx")


if __name__ == "__main__":
    parse_dws_data()
